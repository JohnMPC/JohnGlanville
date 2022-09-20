#retrieve stitch count from all embrodiery files in folder
#create excel spreadsheet with 2 columns. filename and stitch count

from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from itertools import islice
import os
from os.path import  join

#create workbook
wb = Workbook()

#load exsiting workbook
wb = load_workbook("C:/Users/UserPC/Projects/PythonProjects/StitchCount/Stitch_Count.xlsx")

#create an active worksheet
ws = wb.worksheets[0]



#file path for folder with files in 
mypath = 'C:/Users/UserPC/Projects/PythonProjects/StitchCount/'

#new empty lists
data = []
filenamelist=[]
stitchcountlist=[]

#loop through files
for filename in os.listdir(mypath):
    # find files ending with .DST
    if filename.endswith('.DST'):
        # join files ending with .DST to the directory path
        design = join(mypath,filename)
        #open files in read mode and ignore errors as there are lots of special charactors and unsure of encoder.(Data needed is in line 2 and in ABC123 Character set)
        with open(design, mode = 'r', encoding='UTF-8', errors = "ignore") as text:
           #find line 2 in each file
            for line in islice(text,1, 2):
                #for each file add the filename - the last 4 characters and the 2nd line - the first 5 characters to a list.
                filenamelist.append(filename[0:-4])
                stitchcountlist.append(line[5:-1])
        
data = list(zip(filenamelist,stitchcountlist))




row_number = 1
your_column = 1
for i, value in enumerate(filenamelist, start=row_number):
    ws.cell(row=i, column=your_column).value = value

row_number = 1
your_column = 2
for i, value in enumerate(stitchcountlist, start=row_number):
    ws.cell(row=i, column=your_column).value = value


wb.save("C:/Users/UserPC/Projects/PythonProjects/StitchCount/Stitch_Count.xlsx")



               
            